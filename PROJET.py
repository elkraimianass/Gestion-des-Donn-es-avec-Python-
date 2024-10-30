from tkinter import *
from tkinter import ttk ,  messagebox 
import openpyxl   
from openpyxl import Workbook
import pathlib
import pandas 
import matplotlib.pyplot as plt 
import numpy as nb

#CREATION DU FORMULAIRE 

class Formulaire :
	# 1 : FONCTION DU FORMULAIRE
	def __init__(self,fenetre):

		self.fenetre = fenetre
		self.fenetre.title('Formulaire')
		self.fenetre.geometry('900x620')
		self.fenetre.resizable(width=False,height=False)
		self.fenetre.config(bg="#F5F5F5")

		self.photo = PhotoImage(file='C:\\finale\\img4.png')
		label=Label(fenetre,image=self.photo)
		label.place(x=120,y=180)

		texte= Label(fenetre,text='LES FILIERES DES ETUDIANTS DE LA \n BIBLIOTHEQUE DE L\'ENSAO',font=('Verdana',20),fg='black',bg='#F5F5F5').place(x=210,y=10)

		text_nom= Label(fenetre,text='Nom',fg='#000000',bg='#F5F5F5',font=('Verdana',10)).place(x=500,y=100)
		Nom=StringVar()
		self.ecr_nom=Entry(fenetre,textvariable=Nom)
		self.ecr_nom.place(x=500,y=130,width=155,height=35) 

		text_prenom= Label(fenetre,text='Prenom',fg='#000000',bg='#F5F5F5',font=('Verdana',10)).place(x=680,y=100)
		Prenom=StringVar()
		self.ecr_prenom=Entry(fenetre, textvariable=Prenom)
		self.ecr_prenom.place(x=680,y=130,width=155,height=35)

		text_email_aca= Label(fenetre,text='E-mail Academique',fg='#000000',bg='#F5F5F5',font=('Verdana',10)).place(x=500,y=180)
		Email=StringVar()
		self.ecr_emailaca=Entry(fenetre,textvariable=Email)
		self.ecr_emailaca.place(x=500,y=210,width=330,height=35)

		text_mtp= Label(fenetre,text='Mot de Passe',fg='#000000',bg='#F5F5F5',font=('Verdana',10)).place(x=500,y=260)
		mtp=StringVar()
		self.ecr_mtp =Entry(fenetre, show="*" ,textvariable=mtp)
		self.ecr_mtp.place(x=500,y=290,width=330,height=35)

		text_cnferm_mtp= Label(fenetre,text="Confirmer",fg='#000000',bg='#F5F5F5',font=('Verdana',10)).place(x=500,y=340)
		cmtp=StringVar()
		self.ecr_cnfermmtp =Entry(fenetre, show="*" ,textvariable=cmtp)
		self.ecr_cnfermmtp.place(x=500,y=370,width=330,height=35)

		text_fillier= Label(fenetre,text='Filière',fg='#000000',bg='#F5F5F5',font=('Verdana',10)).place(x=500,y=420)
		Filiere=StringVar()
		list1=['SICS','DATA','GINFO','GINDUS','GCIV','GE','ITIRC','GSEIR']
		self.ecr_fillier = ttk.Combobox(fenetre,values=list1)
		self.ecr_fillier.place(x=500,y=450,width=330,height=35)
		self.ecr_fillier.set("SELECTION")

		user_Niveau=Label(fenetre,text='Niveau',fg='#000000',bg='#F5F5F5',font=('Verdana',10)).place(x=500,y=500)
		Niveau=StringVar()
		list2=['3 ème année','4 ème année','5 ème année']
		self.ecr_niv = ttk.Combobox(fenetre,values=list2)
		self.ecr_niv.place(x=500,y=530,width=330,height=35)
		self.ecr_niv.set("SELECTION")

		bouton1=Button(fenetre,text='J\'ENVOIE LES DONNEES',command=self.valide,fg='white',bg='#000000' ,font=("Verdana",7)).place(x=60,y=520,width=150,height=45)
		bouton2=Button(fenetre,text='REINITIALISATION',command=self.rein,fg='white',bg='#000000' ,font=("Verdana",7)).place(x=220,y=520,width=150,height=45)

#VIRIFIER SI LE FICHIER EXCEL EXISTE OU NON 

		fichier = pathlib.Path(r"C:\finale\data.xlsx")
		if fichier.exists():
			pass 
		else :
			fichier = Workbook()
			sheet = fichier.active
			sheet["A1"]="NOM"
			sheet["B1"]="PRENOM"
			sheet["C1"]="EMAIL"
			sheet["D1"]="FILLIERE"
			sheet["E1"]="NIVEAU"
			
			sheet["G1"]="GSEIR"
			sheet["H1"]="GINFO"
			sheet["I1"]="GINDUS"
			sheet["J1"]="SICS"
			sheet["K1"]="DATA"
			sheet["L1"]="ITIRC"
			sheet["M1"]="GCIV"
			sheet["N1"]="GE"


			fichier.save(r"C:\finale\data.xlsx")

    # 2 : FONCTION DE VALIDATION DES DONNEES
	def valide(self):
		nom=self.ecr_nom.get()
		prenom=self.ecr_prenom.get()
		mail=self.ecr_emailaca.get()
		mtpp=self.ecr_mtp.get()
		cmtp=self.ecr_cnfermmtp.get()
		fillier=self.ecr_fillier.get()
		nivv=self.ecr_niv.get()

# VERIFIER SI TOUT LES CHAMPS ONT ETE REMPLIS OU NON 		

		try : 
			if  len(nom) == 0 or len(prenom) == 0 or len(mtpp) == 0 or len(cmtp)==0  or len(mail) == 0 or fillier == 'SELECTION' or nivv== 'SELECTION' :
				raise
		except :
			messagebox.showerror('ERREUR !','VEUILLEZ REMPLIR TOUS LES CHAMPS !')

		else : 
			if  (mtpp != cmtp) :
				messagebox.showerror("ERREUR !","LES DEUX MOTS DE PASS NE SONT PAS IDENTIQUES")
			else : 
				fichier = openpyxl.load_workbook(r"C:\finale\data.xlsx")
				sheet = fichier.active
				sheet.cell(column=1 , row=sheet.max_row+1,value=nom)
				sheet.cell(column=2 , row=sheet.max_row,value=prenom)
				sheet.cell(column=3 , row=sheet.max_row,value=mail)
				sheet.cell(column=4 , row=sheet.max_row,value=fillier)
				sheet.cell(column=5 , row=sheet.max_row,value=nivv)

# STOCKER LES DONNEES DANS LE FICHIER EXCEL  

				fichier.save(r"C:\finale\data.xlsx")
				self.rein()
				messagebox.showinfo(':-)','MERCI D\'AVOIR REMPLIR CE FORMULAIRE  :)')
    
    # 3 : FONCTION DE REINITIALISATION DES DONNEES
	def rein(self) :
		self.ecr_nom.delete(0,END)
		self.ecr_prenom.delete(0 ,END)
		self.ecr_emailaca.delete(0 ,END)
		self.ecr_mtp.delete(0,END)
		self.ecr_cnfermmtp.delete(0,END)
		self.ecr_fillier.set("SELECTION")
		self.ecr_niv.set("SELECTION")

fenetre=Tk()
obj = Formulaire(fenetre)
fenetre.mainloop()

#TRAITEMENT DE FICHIER EXCEL  

print('\n')

liste = pandas.read_excel(r"C:\finale\data.xlsx",usecols="A:E")

var=int(input("POUR AFFICHER LE CONTENU DU FICHIER EXCEL , SAISISSEZ 1 \nPOUR ACCEDER A DES INFORMATIONS SPECIFIQUES SAISISSEZ  2 \n = "))
if var == 1 :
	print(liste)
elif var == 2 :
	print('LES FILIERES SONT  : SICS - GINFO - DATA - GCIV - GINDUS - GSEIR - ITIRC - GE  \n POUR AVOIR LES DONNEES DES ETUDIANTS D\'UNE FILIERE  ENTREZ SON NOM :  ')

	var1=input(" : ")
	def info(var1) :
		if var1 == 'SICS' :
			print(liste[liste.FILLIERE == 'SICS'])
		elif var1 == 'GINFO' :
			print(liste[liste.FILLIERE == 'GINFO'])
		elif var1 == 'DATA' :
			print(liste[liste.FILLIERE == 'DATA'])
		elif var1 == 'GINDUS' :
			print(liste[liste.FILLIERE == 'GINDUS'])
		elif var1 == 'GSIER' :
			print(liste[liste.FILLIERE == 'GSIER'])
		elif var1 == 'ITIRC' :
			print(liste[liste.FILLIERE == 'ITIRC'])
		elif var1 == 'GE' :
			print(liste[liste.FILLIERE == 'GE'])
		else :
			print("DESOLE , LA FILIERE QUE VOUS AVEZ ENTRE N\'EXISTE PAS")

	print(info(var1))
else :
	print("VOUS AVEZ ENTRE AUTRE CHOSE APART 1 OU 2  ") 

print('\n')

#CALCULER LE NOMBRE DES ETUDIANTS DE CHAQUE FILLIER ET STOCKE DANS LA MEME FICHIER EXCEL 

a , b , c , d , e , f , g , h = 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 

for i in liste.FILLIERE :
	if i == 'GSEIR' :
		a += 1	
	elif i == 'GINFO' :
		b += 1
	elif i == 'GINDUS' :
		c += 1 
	elif i =='SICS' :
		d += 1
	elif i =='DATA':
		e += 1 
	elif i =='ITIRC' :
		f += 1 
	elif i == 'GCIV' :
		g += 1 
	elif i =='GE' :
		h += 1

x1 , x2 , x3 , x4 , x5 , x6 , x7 , x8 = a , b , c, d , e , f , g , h 

fichier = openpyxl.load_workbook(r"C:\finale\data.xlsx")
sheet = fichier.active

sheet['G2'].value = x1
sheet['H2'].value = x2
sheet['I2'].value = x3
sheet['J2'].value = x4
sheet['K2'].value = x5
sheet['L2'].value = x6
sheet['M2'].value = x7
sheet['N2'].value = x8

fichier.save(r"C:\finale\data.xlsx")

#POUR AFICHIER LE GRAPHE 

print("POUR AFFICHER LES STATISTIQUES DES FILIERES SAISISSEZ 1 , SINON ENTREZ AUTRE CHOSE POUR FERMER LE PROGRAMME ")
var2=int(input(": "))
if var2 == 1 :
	liste =pandas.read_excel(r"C:\finale\data.xlsx",usecols="G:N")
	GSEIR = liste["GSEIR"]
	GINFO= liste["GINFO"]
	GINDUS = liste["GINDUS"]
	SICS = liste["SICS"]
	DATA= liste["DATA"]
	ITIRC= liste["ITIRC"]
	GCIV = liste["GCIV"]
	GE = liste["GE"]
	labels =("GSEIR","GINFO","GINDUS","SICS","DATA","ITIRC","GCIV","GE")
	x = nb.arange(len(labels))
	y = [GSEIR,GINFO,GINDUS,SICS,DATA,ITIRC,GCIV,GE]
	plt.plot(x, y, color = 'purple', marker = 'o', linestyle = 'dashed')
	plt.ylabel(" NOMBRE D\'ETUDIANTS ",color='purple')
	plt.xlabel("FILLIERES ",color='purple')
	plt.title("STATISTIQUES",color='purple')
	plt.xticks(x,labels)
	plt.show()
	print('MERCI D\'AVOIR ENTRER VOS INFORMATIONS')
elif var2 != 1 :
	print('MERCI D\'AVOIR ENTRER VOS INFORMATIONS')
